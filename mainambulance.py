from tkinter import *
from datetime import date,datetime
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
from openpyxl import*
import openpyxl 
from openpyxl import Workbook
import pathlib

background="#06283D"
framebg="#EDEDED"
framefg="#06283D"


root=Tk()
root.title("Ambulance Booking System")
root.geometry("1250x700+210+100")
root.config(bg=background)


def validate_numeric_input(new_value):
    # Allow only integer input and limit to 10 digits
    if new_value.isdigit() and len(new_value) <= 10:
        return True
    elif new_value == "":
        return True  # Allow clearing the input
    return False

def validate_numeric_input_adhar(new_value):
    # Allow only integer input and limit to 10 digits
    if new_value.isdigit() and len(new_value) <= 12:
        return True
    elif new_value == "":
        return True  # Allow clearing the input
    return False

def validate_numeric_input_pin(new_value):
    # Allow only integer input and limit to 10 digits
    if new_value.isdigit() and len(new_value) <=6:
        return True
    elif new_value == "":
        return True  # Allow clearing the input
    return False
def validate_string_input(new_value):
    # Allow only alphabetic characters and spaces
    if all(char.isalpha() or char.isspace() for char in new_value):
        return True
    return False
vcmd_numeric = (root.register(validate_numeric_input), '%P')
vcmd_numeric_adhar = (root.register(validate_numeric_input_adhar), '%P')
vcmd_numeric_pin = (root.register(validate_numeric_input_pin), '%P')
vcmd_string = (root.register(validate_string_input), '%P')

file=pathlib.Path("Booking_data.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active 
    sheet["A1"]="Registration No."
    sheet["B1"]="Date of Registraton"
    sheet["C1"]="Time of Booking"
    sheet["D1"]="Name of Patient"
    sheet["E1"]="Gender"
    sheet["F1"]="Age of Patient"
    sheet["G1"]="Case of Patient"
    sheet["H1"]="Name of Applicant"
    sheet["I1"]="Mobile No."
    sheet["J1"]="Adhar No."
    sheet["K1"]="Relation"
    sheet["L1"]="State Name"
    sheet["M1"]="City Name"
    sheet["N1"]="Pin Code"
    sheet["O1"]="Hospital Name"
    sheet["P1"]="Alternate No."

    file.save("Booking_data.xlsx")


##Registration 
#now each tiem we have to enter registration no. lets design to automatic enter of reg.no.

def registration_no():
    file=openpyxl.load_workbook('Booking_data.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value
    print(max_row_value)

    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set("1")

#clear
def clear():
     
    Name.set("")
    Case.set("")
    a_name.set("")
    mob.set("")
    Adhar.set("")
    Relation.set("")
    City.set("")
    Pin.set("")
    Hospital.set("")
    Alt_No.set("")
    # gender.set("")
    Age.set("Select Age")
    State.set("Select State")


    registration_no()


    save_button.config(state="normal")

 



#save the data

def Save():
    R1=Registration.get()
    print(R1)
    D1=Date.get()
    print(D1)
    T1= datetime.now().strftime("%I:%M:%S %p")
    print(T1)
    N1=Name.get()
    print(N1)
    try:
        G1=gender
        print(G1)
    except:
        messagebox.showerror("error","Select Gender!")
    A1=Age.get()
    print(A1)
    C1=case_entry.get()
    print(C1)
    AN1=a_name.get()
    print(AN1)
    M1=mob.get()
    print(M1)
    A2=Adhar.get()
    print(A2)
    Re=Relation.get()
    print(Re)
    S1=State.get()
    print(S1)
    C2=City.get()
    print(C2)
    P1=Pin.get()
    print(P1)    
    H1=Hospital.get()
    print(H1)
    A3=Alt_No.get()
    print(A3)
    

    if N1=="" or A1=="Select Age" or C1=="" or AN1=="" or M1=="" or A2=="" or Re=="" or S1=="" or C2=="" or P1=="" or H1=="" or A3=="" :
        messagebox.showerror("error","Few Data is missing!")
    else:
        file=openpyxl.load_workbook("Booking_data.xlsx")
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=D1)
        sheet.cell(column=3,row=sheet.max_row,value=T1)
        sheet.cell(column=4,row=sheet.max_row,value=N1)
        sheet.cell(column=5,row=sheet.max_row,value=G1)
        sheet.cell(column=6,row=sheet.max_row,value=A1)
        sheet.cell(column=7,row=sheet.max_row,value=C1)
        sheet.cell(column=8,row=sheet.max_row,value=AN1)
        sheet.cell(column=9,row=sheet.max_row,value=M1)
        sheet.cell(column=10,row=sheet.max_row,value=A2)
        sheet.cell(column=11,row=sheet.max_row,value=Re)
        sheet.cell(column=12,row=sheet.max_row,value=S1)
        sheet.cell(column=13,row=sheet.max_row,value=C2)
        sheet.cell(column=14,row=sheet.max_row,value=P1)
        sheet.cell(column=15,row=sheet.max_row,value=H1)
        sheet.cell(column=16,row=sheet.max_row,value=A3)


        file.save(r'Booking_data.xlsx')
        
         
        messagebox.showinfo("info","Sucessfully data entered!!!")

        clear()
        registration_no()

#gender
def selection():
    global gender
    value=radio.get()
    if value ==1:
        gender="Male"
    else:
        gender="Female"

#Exit Window
def Exit():
    root.destroy()


#top frame

Label(root,text="email:yescompany@gmail.com",width=5,height=3,bg="#f0687c",anchor="e").pack(side=TOP,fill=X)
Label(root,text="AMBULANCE BOOKING ",width=25,height=2,bg="#C36464",fg='#fff',font='arial 20 bold').pack(side=TOP,fill=X)

#search box to update
Search=StringVar()
Entry(root,textvariable=Search,validate='key',validatecommand=vcmd_numeric,width=10,bd=2,font="arial 16").place(x=900,y=70)
imageicon=PhotoImage(file=r"A:\LANGUAGE\PYHTON\PYTHON\project\1-removebg-preview.png")
Srch=Button(root,text="search",compound=LEFT,image=imageicon,width=123,bg="#68ddfa",font="arial 13 bold")
Srch.place(x=1080,y=66)

imageicon1=PhotoImage(file=r"A:\LANGUAGE\PYHTON\PYTHON\project\Untitled_design-removebg-preview.png")
update_button=Button(root,image=imageicon1,bg="#c36464")
update_button.place(x=110,y=60)

#label of reg. date and time
Label(root,text="Registration No.:",font="arial 13",fg=framebg,bg=background).place(x=30,y=150)
Label(root,text="Date:",font="arial 13",fg=framebg,bg=background).place(x=500,y=150)
Label(root,text="Time:",font="arial 13",fg=framebg,bg=background).place(x=700,y=150)
Registration=IntVar()
Date=StringVar()
Time=StringVar()
#registration number
reg_entry=Entry(root,textvariable=Registration,width=15,font="arial 10").place(x=160,y=150)
registration_no()
#date
today=date.today()
d1=today.strftime("%d/%m/%Y")
date_entry=Entry(root,textvariable=Date,width=15,font="arial 10").place(x=550,y=150)
Date.set(d1)

#update time every one second
def update_time():
    current_time = datetime.now().strftime("%I:%M:%S %p")
    time_label.config(text=current_time)
    root.after(1000, update_time)  # Call itself after 1000ms (1 second)
time_label = Label(root, font="arial 13", fg=framebg, bg=background)
time_label.place(x=750, y=150)
update_time()

# Patient Detail
obj=LabelFrame(root,text="Patient's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=200,relief=GROOVE) 
obj.place(x=30,y=180)

#name of patient
Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Name=StringVar()
name_entry=Entry(obj,textvariable=Name,validate='key',validatecommand=vcmd_string,width=20,font="arial 10")
name_entry.place(x=160,y=50)
 
#gender of patient
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
radio=IntVar()
Radio1=Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
Radio1.place(x=150,y=100)
Radio2=Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
Radio2.place(x=200,y=100)

#age of patient
Label(obj,text="Age:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Age=Combobox(obj,value=['0 to 10','11 to 20','20 to 30','30 to 40','40 to 50','50 to above'],font="Roboto 10",width=17,state="r")
Age.place(x=600,y=50)
Age.set("Select Age")
 
#case of patient
Label(obj,text="Case of Patient:",font="arial 13",bg=framebg,fg=framefg).place(x=450,y=100)
# Case=StringVar()
case_entry=Combobox(obj,value=['Heart Problem','Serious Injury','Cold and flu','Infectious Diseases','Cancer','Pregnancy'],validate='key',validatecommand=vcmd_string,width=20,font="arial 10",state='r')
case_entry.place(x=600,y=100)

#Applicant Details
obj1=LabelFrame(root,text="Applicant's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=300,relief=GROOVE)
obj1.place(x=30,y=400)

#applicant name
Label(obj1,text="Name :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
a_name=StringVar()
a_entry=Entry(obj1,textvariable=a_name,validate='key',validatecommand=vcmd_string,font="arial 10" ).place(x=160,y=50)

#applicant mobile number
Label(obj1,text="Mobile No.:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
mob=StringVar()
mob_entry=Entry(obj1,textvariable=mob,validate='key',validatecommand=vcmd_numeric,width=20,font="arial 10")
mob_entry.place(x=160,y=100)

#applicant adhar number
Label(obj1,text="Adhar No:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Adhar=StringVar()
adhar_entry=Entry(obj1,textvariable=Adhar,validate='key',validatecommand=vcmd_numeric_adhar,width=20,font="arial 10")
adhar_entry.place(x=630,y=50)

#applicant Relation to patient
Label(obj1,text="Relation ",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Relation=StringVar()
relation=Entry(obj1,textvariable=Relation,validate='key',validatecommand=vcmd_string,font="arial 10")
relation.place(x=630,y=100)

#location details
Label(obj1,text="Location Details",font=20 ).place(x=0,y=140)

#state 
Label(obj1,text="State Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=170)
State=Combobox(obj1,value=['Delhi','Mumbai','Uttar Pradesh','Gujarat','Madhya Pradesh','Kerala'],validate='key',validatecommand=vcmd_string,font="Roboto 10",width=17,state="r")
State.place(x=160,y=170)
State.set("Select State")
 
#city
Label(obj1,text="City Name :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=220)
City=StringVar()
city=Entry(obj1,textvariable=City,validate='key',validatecommand=vcmd_string,font="arial 10" ).place(x=160,y=220)
  
#pin code
Label(obj1,text="Pin Code:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=170)
Pin=StringVar()
pin=Entry(obj1,textvariable=Pin,validate='key',validatecommand=vcmd_numeric_pin,font="arial 10" ).place(x=630,y=170)

#create thrid frame
obj2=LabelFrame(root,text="Other Details",font=20,bd=2,width=400,bg=framebg,fg=framefg,height=200,relief=GROOVE)
obj2.place(x=970,y=180)

#hospital name 
Label(obj2,text="Hospital Name :",font="arial 13",bg=framebg,fg=framefg).place(x=20,y=20)
Hospital=StringVar()
Hosp_entry=Entry(obj2,textvariable=Hospital,validate='key',validatecommand=vcmd_string,font="arial 10").place(x=150 ,y=20)

#alternate mobile number 
Label(obj2,text="Alternate No.:",font="arial 13",bg=framebg,fg=framefg).place(x=20,y=60)
Alt_No=StringVar()
alt_entry=Entry(obj2,textvariable=Alt_No,validate='key',validatecommand=vcmd_numeric,font="arial 10").place(x=150 ,y=60)


#save Button 
save_button=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=Save)
save_button.place(x=1000,y=450)
#Reset button
Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="yellow",command=clear).place(x=1000,y=530)

#exit Button
Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="grey",command=Exit).place(x=1000,y=610)


  
root.mainloop()